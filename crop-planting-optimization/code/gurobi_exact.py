# -*- coding: utf-8 -*-
"""
2024年高教社杯数学建模C题
问题1(1)：使用 Gurobi 求解原始模型

说明：
1. 本代码使用 Gurobi 的 addGenConstrMin 表示原始模型中的
   S = min(Q, D)，即超过预期销售量的部分滞销浪费。
2. 为了和前面 SA、GA 的代码保持一致，并降低模型规模，
   这里采用“每个地块每季选择一个主作物并种满该季面积”的模式编码。
3. 输出：
   result1_1_Gurobi_original.xlsx
   gurobi_original_summary.xlsx
"""

from pathlib import Path
import re
import math
import time
import copy
import warnings
from collections import defaultdict

import numpy as np
import pandas as pd
from openpyxl import load_workbook

import gurobipy as gp
from gurobipy import GRB

warnings.filterwarnings("ignore")

# ============================================================
# 0. 全局设置
# ============================================================

SINGLE = "单季"
FIRST = "第一季"
SECOND = "第二季"
SEASONS_OUT = [FIRST, SECOND]

YEARS = list(range(2024, 2031))

GRAIN_NO_RICE = set(range(1, 16))      # 1-15，水稻除外粮食
RICE = 16                              # 16，水稻
VEG1 = set(range(17, 35))              # 17-34，第一季蔬菜
VEG2 = {35, 36, 37}                    # 35-37，第二季蔬菜
MUSHROOM = set(range(38, 42))          # 38-41，食用菌
BEAN = {1, 2, 3, 4, 5, 17, 18, 19}     # 豆类作物

# Gurobi 参数
TIME_LIMIT = 600       # 最长求解时间，单位秒
MIP_GAP = 0.02         # 允许 2% 的相对最优间隙

# 候选模式数量
# 如果运行时报 Model too large，可以把 4 改成 2 或 3
MAX_MODES_PER_LAND = 4
USE_HARD_SPREAD = True
SPREAD_PENALTY = 2e5

# ============================================================
# 1. 路径设置
# ============================================================

BASE_DIR = Path(r"F:\数学建模作业\第六次作业\C题")
TEMPLATE_DIR = BASE_DIR / "附件3"

LAND_FILE = BASE_DIR / "附件1.xlsx"
PLANT_FILE = BASE_DIR / "附件2.xlsx"
TEMPLATE_FILE = TEMPLATE_DIR / "result1_1.xlsx"

OUT_GUROBI = TEMPLATE_DIR / "result1_1_Gurobi_original.xlsx"
OUT_SUMMARY = TEMPLATE_DIR / "gurobi_original_summary.xlsx"

required_files = {
    "附件1": LAND_FILE,
    "附件2": PLANT_FILE,
    "result1_1模板": TEMPLATE_FILE,
}

for name, path in required_files.items():
    if not path.exists():
        raise FileNotFoundError(f"{name} 文件不存在，请检查路径：{path}")

print("路径检查通过：")
print(f"附件1：{LAND_FILE}")
print(f"附件2：{PLANT_FILE}")
print(f"模板文件：{TEMPLATE_FILE}")
print(f"Gurobi输出文件：{OUT_GUROBI}")
print(f"结果汇总：{OUT_SUMMARY}")


# ============================================================
# 2. 工具函数
# ============================================================

def clean_columns(df):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def parse_price(x):
    """把 '2.50-4.00' 或 '2.50~4.00' 转成均值。"""
    if pd.isna(x):
        return np.nan

    if isinstance(x, (int, float, np.integer, np.floating)):
        return float(x)

    s = str(x).strip()
    nums = re.findall(r"\d+(?:\.\d+)?", s)

    if len(nums) >= 2:
        return (float(nums[0]) + float(nums[1])) / 2

    if len(nums) == 1:
        return float(nums[0])

    return np.nan


def normalize_season(x):
    if pd.isna(x):
        return SINGLE

    s = str(x).strip()

    if "二" in s or "2" in s:
        return SECOND

    if "一" in s or "1" in s:
        return FIRST

    return SINGLE


def output_season(season):
    """单季作物在模板中写入第一季。"""
    return FIRST if season in (SINGLE, FIRST) else SECOND


def safe_float(x, default=0.0):
    try:
        if pd.isna(x):
            return default
        return float(x)
    except Exception:
        return default


# ============================================================
# 3. 读取数据
# ============================================================

def read_data():
    land = pd.read_excel(LAND_FILE, sheet_name="乡村的现有耕地")
    land = clean_columns(land)
    land = land[["地块名称", "地块类型", "地块面积/亩"]].dropna(subset=["地块名称"])
    land = land.rename(
        columns={
            "地块名称": "land",
            "地块类型": "land_type",
            "地块面积/亩": "area",
        }
    )
    land["land"] = land["land"].astype(str).str.strip()
    land["land_type"] = land["land_type"].astype(str).str.strip()
    land["area"] = pd.to_numeric(land["area"], errors="coerce")

    crops = pd.read_excel(LAND_FILE, sheet_name="乡村种植的农作物")
    crops = clean_columns(crops)
    crops = crops[["作物编号", "作物名称", "作物类型"]].dropna(subset=["作物编号"])
    crops = crops.rename(
        columns={
            "作物编号": "crop_id",
            "作物名称": "crop_name",
            "作物类型": "crop_type",
        }
    )
    crops["crop_id"] = pd.to_numeric(crops["crop_id"], errors="coerce")
    crops = crops.dropna(subset=["crop_id"])
    crops["crop_id"] = crops["crop_id"].astype(int)
    crops["crop_name"] = crops["crop_name"].astype(str).str.strip()

    crop_name = dict(zip(crops["crop_id"], crops["crop_name"]))
    name_crop = {v: k for k, v in crop_name.items()}

    plant2023 = pd.read_excel(PLANT_FILE, sheet_name="2023年的农作物种植情况")
    plant2023 = clean_columns(plant2023)
    plant2023 = plant2023.rename(
        columns={
            "种植地块": "land",
            "作物编号": "crop_id",
            "作物名称": "crop_name",
            "作物类型": "crop_type",
            "种植面积/亩": "area",
            "种植季次": "season",
        }
    )

    plant2023 = plant2023.dropna(subset=["land", "crop_id"])
    plant2023["land"] = plant2023["land"].astype(str).str.strip()
    plant2023["crop_id"] = pd.to_numeric(plant2023["crop_id"], errors="coerce")
    plant2023 = plant2023.dropna(subset=["crop_id"])
    plant2023["crop_id"] = plant2023["crop_id"].astype(int)
    plant2023["area"] = pd.to_numeric(plant2023["area"], errors="coerce")
    plant2023["season"] = plant2023["season"].apply(normalize_season)

    stats = pd.read_excel(PLANT_FILE, sheet_name="2023年统计的相关数据")
    stats = clean_columns(stats)
    stats = stats.rename(
        columns={
            "作物编号": "crop_id",
            "作物名称": "crop_name",
            "地块类型": "land_type",
            "种植季次": "season",
            "亩产量/斤": "yield",
            "种植成本/(元/亩)": "cost",
            "销售单价/(元/斤)": "price",
        }
    )

    stats = stats.dropna(subset=["crop_id", "land_type", "season"])
    stats["crop_id"] = pd.to_numeric(stats["crop_id"], errors="coerce")
    stats = stats.dropna(subset=["crop_id"])
    stats["crop_id"] = stats["crop_id"].astype(int)
    stats["land_type"] = stats["land_type"].astype(str).str.strip()
    stats["season"] = stats["season"].apply(normalize_season)
    stats["yield"] = pd.to_numeric(stats["yield"], errors="coerce")
    stats["cost"] = pd.to_numeric(stats["cost"], errors="coerce")
    stats["price"] = stats["price"].apply(parse_price)

    return land, crops, crop_name, name_crop, plant2023, stats


land_df, crops_df, CROP_NAME, NAME_CROP, plant2023_df, stats_df = read_data()

LANDS = land_df["land"].tolist()
LAND_TYPE = dict(zip(land_df["land"], land_df["land_type"]))
AREA = dict(zip(land_df["land"], land_df["area"]))
CROP_IDS = sorted(CROP_NAME.keys())

print(f"读取数据完成：地块 {len(LANDS)} 个，作物 {len(CROP_IDS)} 种。")


# ============================================================
# 4. 构造参数
# ============================================================

PARAM = {}

for _, row in stats_df.iterrows():
    key = (
        str(row["land_type"]).strip(),
        normalize_season(row["season"]),
        int(row["crop_id"]),
    )
    PARAM[key] = {
        "yield": safe_float(row["yield"]),
        "cost": safe_float(row["cost"]),
        "price": safe_float(row["price"]),
    }


def get_param(land, season, crop):
    land_type = LAND_TYPE[land]
    season = normalize_season(season)
    crop = int(crop)

    candidates = [
        (land_type, season, crop),
        (land_type, SINGLE, crop),
        (land_type, FIRST, crop),
    ]

    # 智慧大棚第一季蔬菜参数常参考普通大棚第一季
    if land_type == "智慧大棚" and season == FIRST:
        candidates.append(("普通大棚", FIRST, crop))

    for key in candidates:
        if key in PARAM:
            return PARAM[key]

    same_crop = [v for (lt, ss, cc), v in PARAM.items() if cc == crop]

    if same_crop:
        return {
            "yield": float(np.mean([v["yield"] for v in same_crop])),
            "cost": float(np.mean([v["cost"] for v in same_crop])),
            "price": float(np.mean([v["price"] for v in same_crop])),
        }

    raise KeyError(f"找不到参数：地块={land}, 季次={season}, 作物={crop}")


# 2023 年预期销量：用 2023 年同季作物总产量近似
DEMAND = defaultdict(float)

for _, row in plant2023_df.iterrows():
    land = row["land"]
    crop = int(row["crop_id"])
    season = normalize_season(row["season"])
    out_s = output_season(season)
    area = safe_float(row["area"])
    param = get_param(land, season, crop)
    DEMAND[(out_s, crop)] += area * param["yield"]


# 价格：按季次和作物取平均
price_values = defaultdict(list)

for (lt, s, c), val in PARAM.items():
    price_values[(output_season(s), c)].append(val["price"])

PRICE = {k: float(np.mean(v)) for k, v in price_values.items()}


def get_price(out_season, crop):
    if (out_season, crop) in PRICE:
        return PRICE[(out_season, crop)]

    same_crop_prices = [v["price"] for (lt, s, c), v in PARAM.items() if c == crop]

    if same_crop_prices:
        return float(np.mean(same_crop_prices))

    return 0.0


# 2023 年最后一季作物，用于 2024 年重茬衔接
PREV_LAST_2023 = defaultdict(set)
BEAN_AREA_2023 = defaultdict(float)

for land, group in plant2023_df.groupby("land"):
    if any(group["season"] == SECOND):
        last_group = group[group["season"] == SECOND]
    else:
        last_group = group

    for cid in last_group["crop_id"].astype(int):
        PREV_LAST_2023[land].add(int(cid))

    for _, row in group.iterrows():
        cid = int(row["crop_id"])
        if cid in BEAN:
            BEAN_AREA_2023[land] += safe_float(row["area"])


# ============================================================
# 5. 构造每个地块的候选种植模式
# ============================================================

def all_options_for_land(land):
    """返回某地块全部理论可选种植模式。"""
    lt = LAND_TYPE[land]

    if lt in ["平旱地", "梯田", "山坡地"]:
        return [[(FIRST, c)] for c in sorted(GRAIN_NO_RICE)]

    if lt == "水浇地":
        modes = [[(FIRST, RICE)]]
        for c1 in sorted(VEG1):
            for c2 in sorted(VEG2):
                modes.append([(FIRST, c1), (SECOND, c2)])
        return modes

    if lt == "普通大棚":
        modes = []
        for c1 in sorted(VEG1):
            for c2 in sorted(MUSHROOM):
                modes.append([(FIRST, c1), (SECOND, c2)])
        return modes

    if lt == "智慧大棚":
        modes = []
        for c1 in sorted(VEG1):
            for c2 in sorted(VEG1):
                if c1 != c2:
                    modes.append([(FIRST, c1), (SECOND, c2)])
        return modes

    raise ValueError(f"未知地块类型：{lt}")


def mode_first_crops(mode):
    return {c for s, c in mode if s == FIRST}


def mode_last_crops(mode):
    if any(s == SECOND for s, c in mode):
        return {c for s, c in mode if s == SECOND}
    return {c for s, c in mode if s == FIRST}


def mode_has_bean(mode):
    return any(c in BEAN for s, c in mode)


def mode_score(land, mode):
    """用单亩毛利润估计候选模式质量。"""
    score = 0.0
    area = AREA[land]

    for season, crop in mode:
        p = get_param(land, season, crop)
        price = get_price(output_season(season), crop)
        score += area * (price * p["yield"] - p["cost"])

    return score


def unique_modes(modes):
    seen = set()
    result = []

    for mode in modes:
        key = tuple(mode)
        if key not in seen:
            seen.add(key)
            result.append(mode)

    return result


def get_2023_mode_for_land(land):
    """
    提取某地块2023年的实际种植模式。
    若同一季有多个作物，则取面积最大的作物作为主作物。
    """
    group = plant2023_df[plant2023_df["land"] == land]

    if group.empty:
        return None

    mode = []

    first_group = group[group["season"].isin([SINGLE, FIRST])]
    second_group = group[group["season"] == SECOND]

    if not first_group.empty:
        row = first_group.sort_values("area", ascending=False).iloc[0]
        mode.append((FIRST, int(row["crop_id"])))

    if not second_group.empty:
        row = second_group.sort_values("area", ascending=False).iloc[0]
        mode.append((SECOND, int(row["crop_id"])))

    if not mode:
        return None

    return mode


def build_potential_output():
    """
    估计每个作物在每一季的潜在最大供给能力。
    用于修正候选模式评分，避免所有地块都选择同一种高毛利作物。
    """
    potential = defaultdict(float)

    for land in LANDS:
        seen = set()

        for mode in all_options_for_land(land):
            for season, crop in mode:
                out_s = output_season(season)
                key = (out_s, crop)

                if key in seen:
                    continue

                seen.add(key)

                try:
                    p = get_param(land, season, crop)
                    potential[key] += AREA[land] * p["yield"]
                except Exception:
                    pass

    return potential


POTENTIAL_OUTPUT = build_potential_output()


def mode_adjusted_score(land, mode):
    """
    需求修正后的候选模式评分。
    如果某作物潜在供给远大于预期销量，则降低其有效价格，
    避免大量选择该作物造成滞销。
    """
    score = 0.0
    area = AREA[land]

    for season, crop in mode:
        out_s = output_season(season)
        p = get_param(land, season, crop)

        price = get_price(out_s, crop)
        demand = float(DEMAND.get((out_s, crop), 0.0))
        potential = float(POTENTIAL_OUTPUT.get((out_s, crop), 1.0))

        demand_factor = min(1.0, demand / max(potential, 1e-9))

        adjusted_price = price * demand_factor

        score += area * (adjusted_price * p["yield"] - p["cost"])

    return score


def build_candidate_modes():
    """
    每个地块保留若干个候选模式。
    改进点：
    1. 保留2023年实际模式，增强方案多样性；
    2. 保留需求修正后收益最高的模式，减少滞销；
    3. 保留豆类模式，保证三年豆类约束可行；
    4. 保留原始毛利润较高的模式，避免过度压低收益。
    """
    modes_by_land = {}

    for land in LANDS:
        all_modes = all_options_for_land(land)

        # 去掉同一模式内部重茬
        cleaned_modes = []

        for m in all_modes:
            first_set = {c for s, c in m if s == FIRST}
            second_set = {c for s, c in m if s == SECOND}

            if len(first_set & second_set) == 0:
                cleaned_modes.append(m)

        all_modes = cleaned_modes

        ranked_adjusted = sorted(all_modes, key=lambda m: mode_adjusted_score(land, m), reverse=True)
        ranked_gross = sorted(all_modes, key=lambda m: mode_score(land, m), reverse=True)

        selected = []

        def add_mode(m):
            if m is None:
                return

            key = tuple(m)

            if key not in {tuple(x) for x in selected}:
                selected.append(m)

        # 1. 加入2023年实际模式
        mode_2023 = get_2023_mode_for_land(land)

        if mode_2023 is not None and tuple(mode_2023) in {tuple(m) for m in all_modes}:
            add_mode(mode_2023)

        # 2. 加入需求修正后最优模式
        if ranked_adjusted:
            add_mode(ranked_adjusted[0])

        # 3. 加入豆类模式
        bean_modes = [m for m in ranked_adjusted if mode_has_bean(m)]

        if bean_modes:
            add_mode(bean_modes[0])

        # 4. 加入原始毛利润最高模式
        if ranked_gross:
            add_mode(ranked_gross[0])

        # 5. 不足则继续补充需求修正后的高分模式
        for m in ranked_adjusted:
            if len(selected) >= MAX_MODES_PER_LAND:
                break
            add_mode(m)

        # 6. 仍不足则补充原始高收益模式
        for m in ranked_gross:
            if len(selected) >= MAX_MODES_PER_LAND:
                break
            add_mode(m)

        modes_by_land[land] = selected[:MAX_MODES_PER_LAND]

    return modes_by_land


MODES = build_candidate_modes()

print("候选模式构造完成：")
for land in LANDS[:5]:
    print(land, MODES[land])

total_mode_vars = sum(len(MODES[l]) for l in LANDS) * len(YEARS)
print(f"预计模式变量数量：{total_mode_vars}")


# ============================================================
# 6. 建立 Gurobi 原始模型
# ============================================================

model = gp.Model("Q1_1_original_mode_based")

# y[t,l,k] = 1 表示第 t 年地块 l 选择第 k 个候选模式
YKEYS = []

for t in YEARS:
    for land in LANDS:
        for k in range(len(MODES[land])):
            YKEYS.append((t, land, k))

y = model.addVars(YKEYS, vtype=GRB.BINARY, name="y")

# 每个地块每年必须选择一个模式
model.addConstrs(
    (
        gp.quicksum(y[t, land, k] for k in range(len(MODES[land]))) == 1
        for t in YEARS
        for land in LANDS
    ),
    name="choose_one_mode"
)

# 禁止 2024 年与 2023 年最后一季重茬
for land in LANDS:
    prev_crops = PREV_LAST_2023[land]

    for k, mode in enumerate(MODES[land]):
        if len(mode_first_crops(mode) & prev_crops) > 0:
            model.addConstr(
                y[2024, land, k] == 0,
                name=f"no_repeat_2023_{land}_{k}"
            )

# 相邻年份不重茬
for t in YEARS[1:]:
    for land in LANDS:
        relevant_crops = set()

        for k_prev, mode_prev in enumerate(MODES[land]):
            relevant_crops |= mode_last_crops(mode_prev)

        for k_cur, mode_cur in enumerate(MODES[land]):
            relevant_crops |= mode_first_crops(mode_cur)

        for crop in relevant_crops:
            prev_sum = gp.quicksum(
                y[t - 1, land, k]
                for k, mode in enumerate(MODES[land])
                if crop in mode_last_crops(mode)
            )
            cur_sum = gp.quicksum(
                y[t, land, k]
                for k, mode in enumerate(MODES[land])
                if crop in mode_first_crops(mode)
            )

            if prev_sum.size() > 0 and cur_sum.size() > 0:
                model.addConstr(
                    prev_sum + cur_sum <= 1,
                    name=f"no_repeat_year_{t}_{land}_{crop}"
                )

# 三年内至少种一次豆类
for land in LANDS:
    # 2023-2025 窗口
    model.addConstr(
        BEAN_AREA_2023[land]
        + gp.quicksum(
            AREA[land] * y[t, land, k]
            for t in [2024, 2025]
            for k, mode in enumerate(MODES[land])
            if mode_has_bean(mode)
        ) >= AREA[land],
        name=f"bean_2023_2025_{land}"
    )

    # 2024-2030 内部滚动窗口
    for start in range(2024, 2029):
        model.addConstr(
            gp.quicksum(
                AREA[land] * y[t, land, k]
                for t in range(start, start + 3)
                for k, mode in enumerate(MODES[land])
                if mode_has_bean(mode)
            ) >= AREA[land],
            name=f"bean_{start}_{start + 2}_{land}"
        )

# ============================================================
# 作物分散度约束
# ============================================================


# 构造产量键：只为候选模式中出现过的作物建 Q 和 S
QKEYS_SET = set()

for t in YEARS:
    for land in LANDS:
        for mode in MODES[land]:
            for season, crop in mode:
                QKEYS_SET.add((t, output_season(season), crop))

QKEYS = sorted(QKEYS_SET)

Q = model.addVars(QKEYS, lb=0, vtype=GRB.CONTINUOUS, name="Q")
Ssale = model.addVars(QKEYS, lb=0, vtype=GRB.CONTINUOUS, name="Ssale")

# 总产量约束
for t, out_s, crop in QKEYS:
    expr = gp.LinExpr()

    for land in LANDS:
        area = AREA[land]

        for k, mode in enumerate(MODES[land]):
            for season, c in mode:
                if output_season(season) == out_s and c == crop:
                    p = get_param(land, season, crop)
                    expr += area * p["yield"] * y[t, land, k]

    model.addConstr(
        Q[t, out_s, crop] == expr,
        name=f"production_{t}_{out_s}_{crop}"
    )

# 原始模型中的 min 约束：Ssale = min(Q, DEMAND)
for t, out_s, crop in QKEYS:
    demand = float(DEMAND.get((out_s, crop), 0.0))

    model.addGenConstrMin(
        Ssale[t, out_s, crop],
        [Q[t, out_s, crop]],
        constant=demand,
        name=f"sale_min_{t}_{out_s}_{crop}"
    )

# 目标函数：最大化总收入 - 总成本
revenue = gp.quicksum(
    get_price(out_s, crop) * Ssale[t, out_s, crop]
    for t, out_s, crop in QKEYS
)

cost = gp.LinExpr()

for t in YEARS:
    for land in LANDS:
        area = AREA[land]

        for k, mode in enumerate(MODES[land]):
            mode_cost = 0.0

            for season, crop in mode:
                p = get_param(land, season, crop)
                mode_cost += area * p["cost"]

            cost += mode_cost * y[t, land, k]

if USE_HARD_SPREAD:
    model.setObjective(revenue - cost, GRB.MAXIMIZE)
else:
    spread_penalty = SPREAD_PENALTY * gp.quicksum(spread_over.values())
    model.setObjective(revenue - cost - spread_penalty, GRB.MAXIMIZE)

# 求解参数
model.Params.TimeLimit = TIME_LIMIT
model.Params.MIPGap = MIP_GAP

model.update()

print("\n模型建立完成，开始调用 Gurobi 求解...")
print(f"变量数：{model.NumVars}")
print(f"约束数：{model.NumConstrs}")
print(f"一般约束数：{model.NumGenConstrs}")

start_time = time.perf_counter()
model.optimize()
elapsed = time.perf_counter() - start_time


# ============================================================
# 7. 提取结果
# ============================================================

def extract_solution():
    sol = {}

    for t in YEARS:
        sol[t] = {}

        for land in LANDS:
            chosen = None

            for k, mode in enumerate(MODES[land]):
                if y[t, land, k].X > 0.5:
                    chosen = copy.deepcopy(mode)
                    break

            if chosen is None:
                chosen = copy.deepcopy(MODES[land][0])

            sol[t][land] = chosen

    return sol


def evaluate_solution(sol):
    production = defaultdict(float)
    total_cost = 0.0

    for t in YEARS:
        for land in LANDS:
            area = AREA[land]

            for season, crop in sol[t][land]:
                p = get_param(land, season, crop)
                out_s = output_season(season)
                production[(t, out_s, crop)] += area * p["yield"]
                total_cost += area * p["cost"]

    total_revenue = 0.0
    waste = 0.0

    for (t, out_s, crop), q in production.items():
        demand = float(DEMAND.get((out_s, crop), 0.0))
        price = get_price(out_s, crop)

        total_revenue += min(q, demand) * price
        waste += max(0.0, q - demand)

    profit = total_revenue - total_cost

    # 检查重茬、豆类、分散度
    violations = defaultdict(int)

    for land in LANDS:
        # 2024 与 2023 衔接
        if len(mode_first_crops(sol[2024][land]) & PREV_LAST_2023[land]) > 0:
            violations["2024与2023重茬"] += 1

        for t in YEARS[1:]:
            if len(mode_last_crops(sol[t - 1][land]) & mode_first_crops(sol[t][land])) > 0:
                violations["跨年重茬"] += 1

    for land in LANDS:
        windows = [
            (2023, 2025),
            (2024, 2026),
            (2025, 2027),
            (2026, 2028),
            (2027, 2029),
            (2028, 2030),
        ]

        for a, b in windows:
            bean_area = 0.0

            if a == 2023:
                bean_area += BEAN_AREA_2023[land]
                years_in = [y0 for y0 in YEARS if a <= y0 <= b]
            else:
                years_in = list(range(a, b + 1))

            for y0 in years_in:
                if mode_has_bean(sol[y0][land]):
                    bean_area += AREA[land]

            if bean_area + 1e-9 < AREA[land]:
                violations["三年豆类不足"] += 1

    # 分散度统计：这里不作为硬约束，只统计
    spread = defaultdict(set)

    for t in YEARS:
        for land in LANDS:
            for season, crop in sol[t][land]:
                spread[(t, output_season(season), crop)].add(land)

    # 以 2023 年分布地块数的 1.2 倍作为参考上限
    base_spread = defaultdict(set)

    for _, row in plant2023_df.iterrows():
        base_spread[(output_season(row["season"]), int(row["crop_id"]))].add(row["land"])

    for (t, out_s, crop), lands in spread.items():
        base = len(base_spread.get((out_s, crop), set()))
        limit = max(1, math.ceil(1.2 * base)) if base > 0 else 5

        if len(lands) > limit:
            violations["作物分散度超限"] += len(lands) - limit

    return {
        "revenue": total_revenue,
        "cost": total_cost,
        "profit": profit,
        "waste": waste,
        "violations": dict(violations),
    }


# ============================================================
# 8. 写入 Excel 模板
# ============================================================

def write_solution_to_template(sol, output_file):
    wb = load_workbook(TEMPLATE_FILE)

    for t in YEARS:
        ws = wb[str(t)]

        crop_col = {}

        for col in range(3, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value

            if val is None:
                continue

            name = str(val).strip()

            if name in NAME_CROP:
                crop_col[NAME_CROP[name]] = col

        row_map = {}

        for row in range(2, ws.max_row + 1):
            land_name = ws.cell(row=row, column=2).value

            if land_name is None:
                continue

            land_name = str(land_name).strip()

            if land_name not in LAND_TYPE:
                continue

            if row <= 55:
                season = FIRST
            else:
                season = SECOND

            row_map[(season, land_name)] = row

            for col in range(3, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None

        for land in LANDS:
            area = AREA[land]

            for season, crop in sol[t][land]:
                out_s = output_season(season)
                row = row_map.get((out_s, land), None)
                col = crop_col.get(int(crop), None)

                if row is not None and col is not None:
                    ws.cell(row=row, column=col).value = float(area)

    wb.save(output_file)


# ============================================================
# 9. 输出汇总结果
# ============================================================

status_map = {
    GRB.OPTIMAL: "OPTIMAL",
    GRB.TIME_LIMIT: "TIME_LIMIT",
    GRB.INFEASIBLE: "INFEASIBLE",
    GRB.INF_OR_UNBD: "INF_OR_UNBD",
    GRB.UNBOUNDED: "UNBOUNDED",
}

status_name = status_map.get(model.Status, str(model.Status))

if model.SolCount > 0:
    sol = extract_solution()
    detail = evaluate_solution(sol)

    write_solution_to_template(sol, OUT_GUROBI)

    gap_value = None
    try:
        gap_value = model.MIPGap
    except Exception:
        gap_value = None

    summary = pd.DataFrame(
        [
            {
                "算法": "Gurobi-原始模型",
                "模型状态": status_name,
                "运行时间/s": round(elapsed, 4),
                "目标函数值/元": round(model.ObjVal, 2),
                "重新计算利润/元": round(detail["profit"], 2),
                "总收入/元": round(detail["revenue"], 2),
                "总成本/元": round(detail["cost"], 2),
                "滞销浪费产量/斤": round(detail["waste"], 2),
                "MIPGap": gap_value,
                "是否得到可行解": "是",
                "违反约束统计": str(detail["violations"]),
                "候选模式数上限": MAX_MODES_PER_LAND,
            }
        ]
    )

    summary.to_excel(OUT_SUMMARY, index=False)

    print("\n===== Gurobi 求解结果 =====")
    print(summary.to_string(index=False))

    print("\n输出完成：")
    print(OUT_GUROBI)
    print(OUT_SUMMARY)

else:
    print("\nGurobi 未找到可行解。")
    print(f"模型状态：{status_name}")

    if model.Status == GRB.INFEASIBLE:
        print("模型不可行，正在计算 IIS...")
        model.computeIIS()
        iis_file = TEMPLATE_DIR / "gurobi_original_infeasible.ilp"
        model.write(str(iis_file))
        print(f"IIS 文件已输出：{iis_file}")
