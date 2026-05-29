# -*- coding: utf-8 -*-
"""
2024年高教社杯数学建模C题：问题1(1) 原始模型的两种智能优化算法
算法：模拟退火 SA + 遗传算法 GA

运行前请安装：
pip install pandas openpyxl numpy
"""

from pathlib import Path
import re
import math
import time
import random
import copy
import warnings
from collections import defaultdict

import numpy as np
import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

# ============================================================
# 0. 全局常量：一定要放在 normalize_season 函数之前
# ============================================================

SINGLE = "单季"
FIRST = "第一季"
SECOND = "第二季"

YEARS = list(range(2024, 2031))

# 作物集合：按照附件作物编号
GRAIN_NO_RICE = set(range(1, 16))     # 1-15：粮食类作物，水稻除外
RICE = 16                             # 16：水稻
VEG1 = set(range(17, 35))             # 17-34：第一季蔬菜
VEG2 = {35, 36, 37}                   # 35-37：第二季三种蔬菜
MUSHROOM = set(range(38, 42))         # 38-41：食用菌
BEAN = {1, 2, 3, 4, 5, 17, 18, 19}    # 豆类作物

RANDOM_SEED = 2024
random.seed(RANDOM_SEED)
np.random.seed(RANDOM_SEED)

# 算法参数：先用这个规模跑通；要更好结果可以调大
SA_MAX_ITER = 800
GA_POP_SIZE = 30
GA_GENERATIONS = 20

# ============================================================
# 1. 路径设置：你的文件结构是 C题/附件1.xlsx、附件2.xlsx、附件3/result1_1.xlsx
# ============================================================

BASE_DIR = Path(r"F:\数学建模作业\第六次作业\C题")
TEMPLATE_DIR = BASE_DIR / "附件3"

LAND_FILE = BASE_DIR / "附件1.xlsx"
PLANT_FILE = BASE_DIR / "附件2.xlsx"
TEMPLATE_FILE = TEMPLATE_DIR / "result1_1.xlsx"

OUT_SA = TEMPLATE_DIR / "result1_1_SA.xlsx"
OUT_GA = TEMPLATE_DIR / "result1_1_GA.xlsx"
OUT_SUMMARY = TEMPLATE_DIR / "algorithm_summary.xlsx"

required_files = {
    "附件1": LAND_FILE,
    "附件2": PLANT_FILE,
    "result1_1模板": TEMPLATE_FILE,
}

for name, path in required_files.items():
    if not path.exists():
        raise FileNotFoundError(f"{name} 文件不存在，请检查路径：{path}")

print("路径检查通过：")
print(f"附件1：{LAND_FILE}")
print(f"附件2：{PLANT_FILE}")
print(f"模板文件：{TEMPLATE_FILE}")
print(f"SA输出文件：{OUT_SA}")
print(f"GA输出文件：{OUT_GA}")
print(f"算法汇总：{OUT_SUMMARY}")

# ============================================================
# 2. 基础工具函数
# ============================================================

def clean_columns(df):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def parse_price(x):
    """把 Excel 中的 '2.50-4.00' 或 '2.50~4.00' 转成均值。"""
    if pd.isna(x):
        return np.nan

    if isinstance(x, (int, float, np.integer, np.floating)):
        return float(x)

    s = str(x).strip()
    nums = re.findall(r"\d+(?:\.\d+)?", s)

    if len(nums) >= 2:
        return (float(nums[0]) + float(nums[1])) / 2

    if len(nums) == 1:
        return float(nums[0])

    return np.nan


def normalize_season(x):
    """统一季次写法。这里之前报错就是因为 SINGLE 没有提前定义。"""
    if pd.isna(x):
        return SINGLE

    s = str(x).strip()

    if "二" in s or "2" in s:
        return SECOND

    if "一" in s or "1" in s:
        return FIRST

    return SINGLE


def output_season(season):
    """模板要求：单季作物填在第一季。"""
    return FIRST if season in (SINGLE, FIRST) else SECOND


def safe_float(x, default=0.0):
    try:
        if pd.isna(x):
            return default
        return float(x)
    except Exception:
        return default


# ============================================================
# 3. 读取数据
# ============================================================

def read_data():
    # 附件1：地块
    land = pd.read_excel(LAND_FILE, sheet_name="乡村的现有耕地")
    land = clean_columns(land)

    land = land[["地块名称", "地块类型", "地块面积/亩"]].dropna(subset=["地块名称"])
    land = land.rename(
        columns={
            "地块名称": "land",
            "地块类型": "land_type",
            "地块面积/亩": "area",
        }
    )
    land["land"] = land["land"].astype(str).str.strip()
    land["land_type"] = land["land_type"].astype(str).str.strip()
    land["area"] = pd.to_numeric(land["area"], errors="coerce")

    # 附件1：作物
    crops = pd.read_excel(LAND_FILE, sheet_name="乡村种植的农作物")
    crops = clean_columns(crops)

    crops = crops[["作物编号", "作物名称", "作物类型"]].dropna(subset=["作物编号"])
    crops = crops.rename(
        columns={
            "作物编号": "crop_id",
            "作物名称": "crop_name",
            "作物类型": "crop_type",
        }
    )

    crops["crop_id"] = pd.to_numeric(crops["crop_id"], errors="coerce")
    crops = crops.dropna(subset=["crop_id"])
    crops["crop_id"] = crops["crop_id"].astype(int)
    crops["crop_name"] = crops["crop_name"].astype(str).str.strip()

    crop_name = dict(zip(crops["crop_id"], crops["crop_name"]))
    name_crop = {v: k for k, v in crop_name.items()}

    # 附件2：2023年种植情况
    plant2023 = pd.read_excel(PLANT_FILE, sheet_name="2023年的农作物种植情况")
    plant2023 = clean_columns(plant2023)

    plant2023 = plant2023.rename(
        columns={
            "种植地块": "land",
            "作物编号": "crop_id",
            "作物名称": "crop_name",
            "作物类型": "crop_type",
            "种植面积/亩": "area",
            "种植季次": "season",
        }
    )

    plant2023 = plant2023.dropna(subset=["land", "crop_id"])
    plant2023["land"] = plant2023["land"].astype(str).str.strip()
    plant2023["crop_id"] = pd.to_numeric(plant2023["crop_id"], errors="coerce")
    plant2023 = plant2023.dropna(subset=["crop_id"])
    plant2023["crop_id"] = plant2023["crop_id"].astype(int)
    plant2023["area"] = pd.to_numeric(plant2023["area"], errors="coerce")
    plant2023["season"] = plant2023["season"].apply(normalize_season)

    # 附件2：2023年统计数据
    stats = pd.read_excel(PLANT_FILE, sheet_name="2023年统计的相关数据")
    stats = clean_columns(stats)

    stats = stats.rename(
        columns={
            "作物编号": "crop_id",
            "作物名称": "crop_name",
            "地块类型": "land_type",
            "种植季次": "season",
            "亩产量/斤": "yield",
            "种植成本/(元/亩)": "cost",
            "销售单价/(元/斤)": "price",
        }
    )

    stats = stats.dropna(subset=["crop_id", "land_type", "season"])
    stats["crop_id"] = pd.to_numeric(stats["crop_id"], errors="coerce")
    stats = stats.dropna(subset=["crop_id"])
    stats["crop_id"] = stats["crop_id"].astype(int)
    stats["land_type"] = stats["land_type"].astype(str).str.strip()
    stats["season"] = stats["season"].apply(normalize_season)
    stats["yield"] = pd.to_numeric(stats["yield"], errors="coerce")
    stats["cost"] = pd.to_numeric(stats["cost"], errors="coerce")
    stats["price"] = stats["price"].apply(parse_price)

    return land, crops, crop_name, name_crop, plant2023, stats


land_df, crops_df, CROP_NAME, NAME_CROP, plant2023_df, stats_df = read_data()

LANDS = land_df["land"].tolist()
LAND_TYPE = dict(zip(land_df["land"], land_df["land_type"]))
AREA = dict(zip(land_df["land"], land_df["area"]))

print(f"读取数据完成：地块 {len(LANDS)} 个，作物 {len(CROP_NAME)} 种。")


# ============================================================
# 4. 构造参数字典
# ============================================================

PARAM = {}

for _, row in stats_df.iterrows():
    key = (str(row["land_type"]).strip(), normalize_season(row["season"]), int(row["crop_id"]))
    PARAM[key] = {
        "yield": safe_float(row["yield"]),
        "cost": safe_float(row["cost"]),
        "price": safe_float(row["price"]),
    }


def get_param(land, season, crop):
    """获取某地块、季次、作物对应的亩产量、成本、价格。"""
    land_type = LAND_TYPE[land]
    season = normalize_season(season)
    crop = int(crop)

    candidates = [
        (land_type, season, crop),
        (land_type, SINGLE, crop),
        (land_type, FIRST, crop),
    ]

    # 智慧大棚第一季参数在附件中通常没有单列，可参考普通大棚第一季蔬菜参数
    if land_type == "智慧大棚" and season == FIRST:
        candidates.append(("普通大棚", FIRST, crop))

    for key in candidates:
        if key in PARAM:
            return PARAM[key]

    # 若仍找不到，则用同一作物的平均参数兜底，避免程序中断
    same_crop = [v for (lt, ss, cc), v in PARAM.items() if cc == crop]

    if same_crop:
        return {
            "yield": float(np.mean([v["yield"] for v in same_crop])),
            "cost": float(np.mean([v["cost"] for v in same_crop])),
            "price": float(np.mean([v["price"] for v in same_crop])),
        }

    raise KeyError(f"找不到参数：地块={land}, 地块类型={land_type}, 季次={season}, 作物={crop}")


def allowed_options(land):
    """
    返回某个地块每年允许的种植模式。
    为了便于智能算法编码，这里每个季次只选一个主作物并种满该季面积。
    """
    lt = LAND_TYPE[land]

    if lt in ["平旱地", "梯田", "山坡地"]:
        return [[(SINGLE, c)] for c in sorted(GRAIN_NO_RICE)]

    if lt == "水浇地":
        modes = [[(SINGLE, RICE)]]

        for c1 in sorted(VEG1):
            for c2 in sorted(VEG2):
                modes.append([(FIRST, c1), (SECOND, c2)])

        return modes

    if lt == "普通大棚":
        modes = []

        for c1 in sorted(VEG1):
            for c2 in sorted(MUSHROOM):
                modes.append([(FIRST, c1), (SECOND, c2)])

        return modes

    if lt == "智慧大棚":
        modes = []

        for c1 in sorted(VEG1):
            for c2 in sorted(VEG1):
                if c1 != c2:
                    modes.append([(FIRST, c1), (SECOND, c2)])

        return modes

    raise ValueError(f"未知地块类型：{lt}")


OPTIONS = {land: allowed_options(land) for land in LANDS}


def first_crops(mode):
    return {c for s, c in mode if s in (SINGLE, FIRST)}


def last_crops(mode):
    if any(s == SECOND for s, c in mode):
        return {c for s, c in mode if s == SECOND}
    return {c for s, c in mode if s in (SINGLE, FIRST)}


def has_bean(mode):
    return any(c in BEAN for s, c in mode)


def mode_conflicts_prev(mode, prev_last):
    return len(first_crops(mode) & set(prev_last)) > 0


# ============================================================
# 5. 构造 2023 年基准参数：需求、重茬衔接、豆类面积
# ============================================================

PREV_LAST_2023 = defaultdict(set)
BEAN_AREA_2023 = defaultdict(float)

for land, group in plant2023_df.groupby("land"):
    # 2023年最后一季作物，用于2024年重茬判断
    if any(group["season"] == SECOND):
        last_group = group[group["season"] == SECOND]
    else:
        last_group = group

    for cid in last_group["crop_id"].astype(int):
        PREV_LAST_2023[land].add(int(cid))

    # 2023年豆类面积
    for _, row in group.iterrows():
        cid = int(row["crop_id"])

        if cid in BEAN:
            BEAN_AREA_2023[land] += safe_float(row["area"])


# 预期销售量：用 2023 年同季作物总产量近似
DEMAND = defaultdict(float)

for _, row in plant2023_df.iterrows():
    land = row["land"]
    crop = int(row["crop_id"])
    season = normalize_season(row["season"])
    out_s = output_season(season)
    area = safe_float(row["area"])
    param = get_param(land, season, crop)

    DEMAND[(out_s, crop)] += area * param["yield"]


# 销售价格：按季次和作物取平均
price_values = defaultdict(list)

for (lt, s, c), val in PARAM.items():
    price_values[(output_season(s), c)].append(val["price"])

PRICE = {k: float(np.mean(v)) for k, v in price_values.items()}


def get_price(out_season, crop):
    if (out_season, crop) in PRICE:
        return PRICE[(out_season, crop)]

    same_crop_prices = [v["price"] for (lt, s, c), v in PARAM.items() if c == crop]

    if same_crop_prices:
        return float(np.mean(same_crop_prices))

    return 0.0


# 分散度上限：根据2023年分布情况设置；未出现作物给一个较宽松上限
base_spread = defaultdict(set)

for _, row in plant2023_df.iterrows():
    base_spread[(output_season(row["season"]), int(row["crop_id"]))].add(row["land"])

MAX_SPREAD = {}

for s in [FIRST, SECOND]:
    for crop in CROP_NAME:
        cnt = len(base_spread.get((s, crop), set()))

        if cnt > 0:
            MAX_SPREAD[(s, crop)] = max(1, math.ceil(1.2 * cnt))
        else:
            MAX_SPREAD[(s, crop)] = 5


# ============================================================
# 6. 解的生成、修复与评价
# ============================================================

def choose_mode(land, prev_last=None, prefer_bean=False):
    """在某地块可选模式中随机选一个，尽量避免与上一季重茬。"""
    opts = OPTIONS[land]
    candidates = opts

    if prev_last is not None:
        candidates = [m for m in candidates if not mode_conflicts_prev(m, prev_last)]

    if prefer_bean:
        bean_candidates = [m for m in candidates if has_bean(m)]

        if bean_candidates:
            candidates = bean_candidates

    if not candidates:
        candidates = opts

    return copy.deepcopy(random.choice(candidates))


def random_solution():
    """顺序生成一个初始解，尽量避免跨年重茬。"""
    sol = {}

    for y in YEARS:
        sol[y] = {}

        for land in LANDS:
            if y == 2024:
                prev_last = PREV_LAST_2023[land]
            else:
                prev_last = last_crops(sol[y - 1][land])

            sol[y][land] = choose_mode(land, prev_last=prev_last)

    return sol


def repair_solution(sol, rounds=3):
    """修复重茬和三年豆类约束。"""
    sol = copy.deepcopy(sol)

    for _ in range(rounds):
        # 修复跨年重茬
        for y in YEARS:
            for land in LANDS:
                if y == 2024:
                    prev_last = PREV_LAST_2023[land]
                else:
                    prev_last = last_crops(sol[y - 1][land])

                if mode_conflicts_prev(sol[y][land], prev_last):
                    sol[y][land] = choose_mode(land, prev_last=prev_last)

        # 修复三年豆类约束
        for land in LANDS:
            windows = [
                (2023, 2025),
                (2024, 2026),
                (2025, 2027),
                (2026, 2028),
                (2027, 2029),
                (2028, 2030),
            ]

            for a, b in windows:
                bean_area = 0.0

                if a == 2023:
                    bean_area += BEAN_AREA_2023[land]
                    years_in = [y for y in YEARS if a <= y <= b]
                else:
                    years_in = list(range(a, b + 1))

                for y in years_in:
                    if has_bean(sol[y][land]):
                        bean_area += AREA[land]

                if bean_area + 1e-9 < AREA[land]:
                    # 优先选择窗口中间年份种豆
                    for y in sorted(years_in, key=lambda yy: abs(yy - sum(years_in) / len(years_in))):
                        if y == 2024:
                            prev_last = PREV_LAST_2023[land]
                        else:
                            prev_last = last_crops(sol[y - 1][land])

                        bean_modes = [
                            m for m in OPTIONS[land]
                            if has_bean(m) and not mode_conflicts_prev(m, prev_last)
                        ]

                        if bean_modes:
                            sol[y][land] = copy.deepcopy(random.choice(bean_modes))
                            break

    return sol


def flatten_records(sol):
    records = []

    for y in YEARS:
        for land in LANDS:
            for season, crop in sol[y][land]:
                records.append(
                    {
                        "year": y,
                        "land": land,
                        "land_type": LAND_TYPE[land],
                        "season": season,
                        "out_season": output_season(season),
                        "crop": int(crop),
                        "crop_name": CROP_NAME[int(crop)],
                        "area": float(AREA[land]),
                    }
                )

    return records


def evaluate(sol, return_detail=False):
    """
    原始模型目标函数：
    收入 = price * min(总产量, 预期销售量)
    超过预期销售量的产量直接滞销浪费，不产生收入。
    """
    records = flatten_records(sol)

    production = defaultdict(float)
    cost = 0.0

    for r in records:
        param = get_param(r["land"], r["season"], r["crop"])
        production[(r["year"], r["out_season"], r["crop"])] += r["area"] * param["yield"]
        cost += r["area"] * param["cost"]

    revenue = 0.0
    waste = 0.0

    for (y, s, crop), q in production.items():
        demand = DEMAND.get((s, crop), 0.0)
        price = get_price(s, crop)

        revenue += min(q, demand) * price
        waste += max(0.0, q - demand)

    profit = revenue - cost

    penalty = 0.0
    violations = defaultdict(int)

    # 硬约束：跨年重茬
    for y in YEARS:
        for land in LANDS:
            if y == 2024:
                prev_last = PREV_LAST_2023[land]
            else:
                prev_last = last_crops(sol[y - 1][land])

            if mode_conflicts_prev(sol[y][land], prev_last):
                penalty += 1e8
                violations["跨年重茬"] += 1

    # 硬约束：同年两季重茬
    for y in YEARS:
        for land in LANDS:
            mode = sol[y][land]

            if len(mode) == 2 and mode[0][1] == mode[1][1]:
                penalty += 1e8
                violations["同年两季重茬"] += 1

    # 硬约束：三年豆类
    for land in LANDS:
        windows = [
            (2023, 2025),
            (2024, 2026),
            (2025, 2027),
            (2026, 2028),
            (2027, 2029),
            (2028, 2030),
        ]

        for a, b in windows:
            bean_area = 0.0

            if a == 2023:
                bean_area += BEAN_AREA_2023[land]
                years_in = [y for y in YEARS if a <= y <= b]
            else:
                years_in = list(range(a, b + 1))

            for y in years_in:
                if has_bean(sol[y][land]):
                    bean_area += AREA[land]

            if bean_area + 1e-9 < AREA[land]:
                penalty += 1e8
                violations["三年豆类不足"] += 1

    # 软约束：分散度
    spread = defaultdict(set)

    for r in records:
        spread[(r["year"], r["out_season"], r["crop"])].add(r["land"])

    for (y, s, crop), land_set in spread.items():
        limit = MAX_SPREAD.get((s, crop), 5)
        over = len(land_set) - limit

        if over > 0:
            penalty += 2e5 * over
            violations["作物分散度超限"] += over

    fitness = profit - penalty

    detail = {
        "revenue": revenue,
        "cost": cost,
        "profit": profit,
        "waste_jin": waste,
        "penalty": penalty,
        "violations": dict(violations),
    }

    if return_detail:
        return fitness, profit, penalty, detail

    return fitness


# ============================================================
# 7. 模拟退火算法 SA
# ============================================================

def mutate_solution(sol, n_changes=1):
    new_sol = copy.deepcopy(sol)

    for _ in range(n_changes):
        y = random.choice(YEARS)
        land = random.choice(LANDS)

        if y == 2024:
            prev_last = PREV_LAST_2023[land]
        else:
            prev_last = last_crops(new_sol[y - 1][land])

        new_sol[y][land] = choose_mode(land, prev_last=prev_last)

    return new_sol


def simulated_annealing(max_iter=SA_MAX_ITER, T0=1e7, alpha=0.995):
    start = time.perf_counter()

    current = repair_solution(random_solution())
    current_fit = evaluate(current)

    best = copy.deepcopy(current)
    best_fit = current_fit

    temp = T0

    for it in range(1, max_iter + 1):
        cand = mutate_solution(current, n_changes=random.randint(1, 3))

        if it % 30 == 0:
            cand = repair_solution(cand, rounds=1)

        cand_fit = evaluate(cand)
        delta = cand_fit - current_fit

        if delta >= 0 or random.random() < math.exp(delta / max(temp, 1e-9)):
            current = cand
            current_fit = cand_fit

            if cand_fit > best_fit:
                best = copy.deepcopy(cand)
                best_fit = cand_fit

        temp *= alpha

        if it % 500 == 0 or it == max_iter:
            _, p, pen, d = evaluate(best, return_detail=True)
            print(f"[SA] iter={it}, profit={p:,.2f}, penalty={pen:,.2f}, violations={d['violations']}")

    elapsed = time.perf_counter() - start

    return best, elapsed


# ============================================================
# 8. 遗传算法 GA
# ============================================================

def crossover(p1, p2):
    child = copy.deepcopy(p1)

    for y in YEARS:
        for land in LANDS:
            if random.random() < 0.5:
                child[y][land] = copy.deepcopy(p2[y][land])

    return child


def ga_optimize(pop_size=GA_POP_SIZE, generations=GA_GENERATIONS, elite_size=4, mutation_rate=0.15):
    start = time.perf_counter()

    population = [repair_solution(random_solution()) for _ in range(pop_size)]

    best = None
    best_fit = -1e100

    def tournament(scored, k=3):
        chosen = random.sample(scored, min(k, len(scored)))
        chosen.sort(key=lambda item: item[0], reverse=True)
        return chosen[0][1]

    for gen in range(1, generations + 1):
        scored = [(evaluate(ind), ind) for ind in population]
        scored.sort(key=lambda item: item[0], reverse=True)

        if scored[0][0] > best_fit:
            best_fit = scored[0][0]
            best = copy.deepcopy(scored[0][1])

        new_pop = [copy.deepcopy(ind) for _, ind in scored[:elite_size]]

        while len(new_pop) < pop_size:
            p1 = tournament(scored)
            p2 = tournament(scored)

            child = crossover(p1, p2)

            if random.random() < mutation_rate:
                child = mutate_solution(child, n_changes=random.randint(1, 5))

            if gen % 5 == 0:
                child = repair_solution(child, rounds=1)

            new_pop.append(child)

        population = new_pop

        if gen % 10 == 0 or gen == generations:
            _, p, pen, d = evaluate(best, return_detail=True)
            print(f"[GA] gen={gen}, profit={p:,.2f}, penalty={pen:,.2f}, violations={d['violations']}")

    elapsed = time.perf_counter() - start

    return best, elapsed


# ============================================================
# 9. 写入 result1_1.xlsx 模板
# ============================================================

def write_solution_to_template(sol, output_file):
    wb = load_workbook(TEMPLATE_FILE)

    for y in YEARS:
        ws = wb[str(y)]

        # 表头：作物名 -> 列号
        crop_col = {}

        for col in range(3, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value

            if val is None:
                continue

            crop_name = str(val).strip()

            if crop_name in NAME_CROP:
                crop_col[NAME_CROP[crop_name]] = col

        # 地块行：第一季 2-55，第二季 56-83
        row_map = {}

        for row in range(2, ws.max_row + 1):
            land_name = ws.cell(row=row, column=2).value

            if land_name is None:
                continue

            land_name = str(land_name).strip()

            if land_name not in LAND_TYPE:
                continue

            if row <= 55:
                season = FIRST
            else:
                season = SECOND

            row_map[(season, land_name)] = row

            # 清空旧结果
            for col in range(3, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None

        # 写入解
        for land in LANDS:
            for season, crop in sol[y][land]:
                s_out = output_season(season)
                row = row_map.get((s_out, land), None)
                col = crop_col.get(int(crop), None)

                if row is not None and col is not None:
                    ws.cell(row=row, column=col).value = float(AREA[land])

    wb.save(output_file)


# ============================================================
# 10. 汇总结果
# ============================================================

def summarize_solution(name, sol, elapsed):
    fitness, profit, penalty, detail = evaluate(sol, return_detail=True)

    hard_violations = {
        k: v for k, v in detail["violations"].items()
        if k in ["跨年重茬", "同年两季重茬", "三年豆类不足"]
    }

    valid = len(hard_violations) == 0

    return {
        "算法": name,
        "运行时间/s": round(elapsed, 4),
        "适应度=利润-惩罚": round(fitness, 2),
        "原始利润/元": round(profit, 2),
        "总收入/元": round(detail["revenue"], 2),
        "总成本/元": round(detail["cost"], 2),
        "滞销浪费产量/斤": round(detail["waste_jin"], 2),
        "惩罚项": round(penalty, 2),
        "是否得到有效结果": "是" if valid else "否",
        "违反约束统计": str(detail["violations"]),
    }


# ============================================================
# 11. 主程序
# ============================================================

if __name__ == "__main__":
    print("\n开始运行模拟退火算法 SA...")
    sa_sol, sa_time = simulated_annealing()

    print("\n开始运行遗传算法 GA...")
    ga_sol, ga_time = ga_optimize()

    print("\n正在写入 Excel 结果文件...")
    write_solution_to_template(sa_sol, OUT_SA)
    write_solution_to_template(ga_sol, OUT_GA)

    summary = pd.DataFrame(
        [
            summarize_solution("模拟退火 SA", sa_sol, sa_time),
            summarize_solution("遗传算法 GA", ga_sol, ga_time),
        ]
    )

    summary.to_excel(OUT_SUMMARY, index=False)

    print("\n===== 算法结果汇总 =====")
    print(summary.to_string(index=False))

    print("\n输出完成：")
    print(OUT_SA)
    print(OUT_GA)
    print(OUT_SUMMARY)
